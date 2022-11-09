from flask import Flask, jsonify, request
from db import connection, mycursor
import os.path
# import pandas as pd
# import xlrd
from openpyxl import Workbook, load_workbook
import datetime

app = Flask(__name__)
@app.route("/billing-api/health")
def index_test_bd():
    # with connection.cursor() as mycursor:
    #             mycursor = connection.cursor(dictionary=True)
    #             stmt = "select 1"
    #             mycursor.execute(stmt)
    #             connection.commit()
    return jsonify({"message":"billing server health check successful"}), 200


@app.route('/rates', methods=['GET', 'POST'])
def rates():
    if request.method == 'POST':
        # receive post parameters
        body = request.get_json()
        # name = body['name']
        file = body['file']
        print(os.path.isfile(file))
        # check if file exist
        if os.path.isfile(file):
            # df = pd.read_excel(file)
            data = []
            df = load_workbook(file)
            sheet = df.active
            row_count = sheet.max_row
            for rows in sheet.iter_rows():
                row_cells = []
                for cell in rows:
                    row_cells.append(cell.value)
                data.append(tuple(row_cells))
                
            # establish db connection
            with connection.cursor() as mycursor:
                mycursor = connection.cursor(dictionary=True)
                stmt = "INSERT INTO Rates (`product_id`, `rate`, `scope`) VALUES (%s, %s, %s)"
                mycursor.executemany(stmt, data[1:])
                connection.commit()
                return jsonify({"msg": "Rates uploaded successfully!"}), 201
        else:
            return jsonify({"msg": "Unsupported file format!!!"}), 204
    else:
        with connection.cursor() as mycursor:
            # mycursor = connection.cursor(dictionary=True)
            stmt = "SELECT * FROM Rates"
            mycursor.execute(stmt)
            stmt_result = mycursor.fetchall()
            # export data to excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "rates"

            ws.append(['Product', 'Rate', 'Scope'])

            for row in stmt_result:
                ws.append(list(row))
            wb.save(f"in/rates-{datetime.datetime.now().strftime('%Y%m%d')}.xlsx")
            
            return jsonify({"msg": "Rates downloaded successfully!"})


@app.route('/bill/<id>')
def getbill():
    id = request.args.get('id')
    t1 = request.args.get('t1')
    t2 = request.args.get('t2')


@app.errorhandler(500)
def internal_server_error(error):
    return jsonify({
        "success": False,
        "error": 500,
        "msg": "Internal Server Error"
    })

if __name__ == '__main__':
    app.run(debug=True, port=5000)
    